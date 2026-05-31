"""Spreadsheet I/O for the user loader, backed by openpyxl (no pandas).

This is one *input adapter* among potential others. The core logic in
``core.py`` works on plain dict rows, so the planned Flask web UI can feed
rows straight in without touching this module.
"""

from __future__ import annotations

from openpyxl import load_workbook


def read_rows(file_path: str, sheet_name: str, required_columns) -> list[dict]:
    """Read a worksheet into a list of ``{column: str}`` row dicts.

    The first row is treated as the header. Cell values are coerced to
    stripped strings (blank cells become ``""``). Raises if any required
    column is missing.
    """
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a clear message to the user
        raise RuntimeError(f"Can't read spreadsheet '{file_path}': {exc}") from exc

    try:
        worksheet = workbook[sheet_name]
    except KeyError as exc:
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
        ) from exc

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise RuntimeError("Spreadsheet is empty") from None

    columns = [("" if cell is None else str(cell).strip()) for cell in header]

    missing = [c for c in required_columns if c not in columns]
    if missing:
        raise RuntimeError(
            f"Missing required column(s): {missing}. Expected: {sorted(required_columns)}"
        )

    rows: list[dict] = []
    for raw in rows_iter:
        row = {
            col: ("" if value is None else str(value).strip())
            for col, value in zip(columns, raw)
        }
        rows.append(row)

    workbook.close()
    return rows
